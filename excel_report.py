# -*- coding: utf-8 -*-
"""
excel_report.py — Genera Excel profesional con openpyxl.
6 hojas: Resumen, Datos Completos, Top Vendidos, Top Rentabilidad,
         Margen Negativo, Analisis Pareto.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ---------------------------------------------------------------------------
# Paleta de colores
# ---------------------------------------------------------------------------
COLOR_HEADER      = '1B3A5C'   # azul oscuro
COLOR_HEADER_FONT = 'FFFFFF'   # blanco
COLOR_ALT_ROW     = 'F8F9FA'   # gris claro alterno
COLOR_DORADO      = 'C8A84E'
COLOR_VERDE       = '27AE60'
COLOR_ROJO        = 'E74C3C'
COLOR_TITULO_BG   = 'D6E4F0'

FMT_MONEDA   = '$ #,##0.00'
FMT_PORCENTAJE = '0.00"%"'
FMT_ENTERO   = '#,##0'
FMT_DECIMAL  = '#,##0.00'


# ---------------------------------------------------------------------------
# Helpers de estilos
# ---------------------------------------------------------------------------

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _border_thin():
    side = Side(style='thin', color='CCCCCC')
    return Border(left=side, right=side, top=side, bottom=side)


def _header_style(ws, row, cols, color=COLOR_HEADER):
    for col in cols:
        cell = ws.cell(row=row, column=col)
        cell.fill = _fill(color)
        cell.font = Font(name='Arial', bold=True, color=COLOR_HEADER_FONT, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _border_thin()


def _data_row_style(ws, row, col_start, col_end, alternate):
    bg = COLOR_ALT_ROW if alternate else 'FFFFFF'
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _fill(bg)
        cell.font = Font(name='Arial', size=9)
        cell.border = _border_thin()


def _autofit(ws, min_width=8, max_width=40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _write_table(ws, headers, rows_data, start_row=1, fmt_map=None, neg_col=None):
    """
    Escribe encabezados + datos con estilos.
    fmt_map: dict {col_index_1based: format_string}
    neg_col: indice de columna (1-based) para colorear negativos en rojo
    Retorna la ultima fila escrita.
    """
    fmt_map = fmt_map or {}
    n_cols = len(headers)

    # Encabezados
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    _header_style(ws, start_row, range(1, n_cols + 1))
    ws.row_dimensions[start_row].height = 30

    # Datos
    for i, row_data in enumerate(rows_data):
        row_num = start_row + 1 + i
        alternate = i % 2 == 1
        _data_row_style(ws, row_num, 1, n_cols, alternate)
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=c, value=val)
            if c in fmt_map:
                cell.number_format = fmt_map[c]
            cell.alignment = Alignment(horizontal='right' if isinstance(val, (int, float)) else 'left',
                                       vertical='center')
            # Colorear negativos
            if neg_col and c == neg_col and isinstance(val, (int, float)) and val < 0:
                cell.font = Font(name='Arial', size=9, color=COLOR_ROJO, bold=True)

    return start_row + len(rows_data)


# ---------------------------------------------------------------------------
# HOJA 1 — Resumen
# ---------------------------------------------------------------------------

def _hoja_resumen(wb, resumen, metadata):
    ws = wb.active
    ws.title = 'Resumen'
    ws.sheet_view.showGridLines = False

    sucursal = resumen.get('sucursal') or metadata.get('sucursal', '')
    fecha_d = resumen.get('fecha_desde') or metadata.get('fecha_desde', '')
    fecha_h = resumen.get('fecha_hasta') or metadata.get('fecha_hasta', '')
    titulo = 'Reporte de Ventas — %s — %s a %s' % (sucursal, fecha_d, fecha_h)

    # Titulo principal
    ws.merge_cells('A1:D1')
    t = ws['A1']
    t.value = titulo
    t.font = Font(name='Arial', bold=True, size=14, color=COLOR_HEADER)
    t.alignment = Alignment(horizontal='center', vertical='center')
    t.fill = _fill(COLOR_TITULO_BG)
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:D2')
    ws['A2'].value = 'Generado con Generador de Reportes — Grupo Petri'
    ws['A2'].font = Font(name='Arial', italic=True, size=9, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Metricas
    metricas = [
        ('Total Ventas (Precio)',    resumen['total_ventas'],       FMT_MONEDA),
        ('Total Costos',             resumen['total_costo'],        FMT_MONEDA),
        ('Rentabilidad Total',       resumen['total_rentabilidad'], FMT_MONEDA),
        ('Margen Global',            resumen['margen_global'],      '0.00"%"'),
        ('Cantidad de Productos',    resumen['cantidad_productos'], FMT_ENTERO),
        ('Productos con ventas > 0', resumen['cantidad_activos'],   FMT_ENTERO),
    ]

    # Encabezados de tabla
    ws.cell(row=4, column=1, value='Metrica').font = Font(name='Arial', bold=True, size=10, color=COLOR_HEADER_FONT)
    ws.cell(row=4, column=1).fill = _fill(COLOR_HEADER)
    ws.cell(row=4, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=4, column=2, value='Valor').font = Font(name='Arial', bold=True, size=10, color=COLOR_HEADER_FONT)
    ws.cell(row=4, column=2).fill = _fill(COLOR_HEADER)
    ws.cell(row=4, column=2).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 24

    for i, (label, val, fmt) in enumerate(metricas):
        row = 5 + i
        alternate = i % 2 == 1
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=val)
        c2.number_format = fmt
        for c in [c1, c2]:
            c.fill = _fill(COLOR_ALT_ROW if alternate else 'FFFFFF')
            c.font = Font(name='Arial', size=10)
            c.border = _border_thin()
        c1.alignment = Alignment(horizontal='left', vertical='center')
        c2.alignment = Alignment(horizontal='right', vertical='center')

    # Highlights
    row = 12
    ws.cell(row=row, column=1, value='Destacados').font = Font(name='Arial', bold=True, size=11, color=COLOR_HEADER)
    ws.cell(row=row, column=1).fill = _fill(COLOR_TITULO_BG)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 22

    highlights = []
    if resumen.get('mas_vendido'):
        mv = resumen['mas_vendido']
        highlights.append(('Producto mas vendido (unidades)',
                            '%s — %s' % (mv['codigo'], mv['descripcion']),
                            '%s un.' % '{:,.2f}'.format(mv['unidades'])))
    if resumen.get('mas_rentable'):
        mr = resumen['mas_rentable']
        highlights.append(('Producto mas rentable ($)',
                            '%s — %s' % (mr['codigo'], mr['descripcion']),
                            '$ %s' % '{:,.2f}'.format(mr['rentabilidad'])))
    if resumen.get('mas_facturado'):
        mf = resumen['mas_facturado']
        highlights.append(('Mayor facturacion ($)',
                            '%s — %s' % (mf['codigo'], mf['descripcion']),
                            '$ %s' % '{:,.2f}'.format(mf['precio'])))

    for i, (lbl, prod, val) in enumerate(highlights):
        r = 13 + i
        ws.cell(row=r, column=1, value=lbl).font = Font(name='Arial', bold=True, size=9)
        ws.cell(row=r, column=2, value=prod).font = Font(name='Arial', size=9)
        ws.cell(row=r, column=3, value=val).font = Font(name='Arial', size=9, color=COLOR_VERDE)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = _border_thin()
            ws.cell(row=r, column=c).fill = _fill('FFFFFF' if i % 2 == 0 else COLOR_ALT_ROW)

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15


# ---------------------------------------------------------------------------
# HOJA 2 — Datos Completos
# ---------------------------------------------------------------------------

def _hoja_datos(wb, df):
    ws = wb.create_sheet('Datos Completos')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    headers = ['Codigo', 'Descripcion', 'Unidades', 'Costo ($)', 'Precio ($)',
               'Rentabilidad ($)', 'Margen (%)', 'Participacion (%)']
    fmt_map = {3: FMT_DECIMAL, 4: FMT_MONEDA, 5: FMT_MONEDA,
               6: FMT_MONEDA, 7: FMT_PORCENTAJE, 8: FMT_PORCENTAJE}

    rows_data = [
        (row['codigo'], row['descripcion'], row['unidades'],
         row['costo'], row['precio'], row['rentabilidad'],
         row['margen'], row['participacion'])
        for _, row in df.iterrows()
    ]
    _write_table(ws, headers, rows_data, fmt_map=fmt_map, neg_col=7)
    _autofit(ws)


# ---------------------------------------------------------------------------
# HOJA 3 — Top Vendidos (+ grafico)
# ---------------------------------------------------------------------------

def _hoja_top(wb, sheet_name, df_top, value_col, value_label, chart_title):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    headers = ['Codigo', 'Descripcion', 'Unidades', 'Costo ($)',
               'Precio ($)', 'Rentabilidad ($)', 'Margen (%)']
    fmt_map = {3: FMT_DECIMAL, 4: FMT_MONEDA, 5: FMT_MONEDA,
               6: FMT_MONEDA, 7: FMT_PORCENTAJE}

    rows_data = [
        (row['codigo'], row['descripcion'], row['unidades'],
         row['costo'], row['precio'], row['rentabilidad'], row['margen'])
        for _, row in df_top.iterrows()
    ]
    last_row = _write_table(ws, headers, rows_data, fmt_map=fmt_map)
    _autofit(ws)

    # --- Grafico de barras ---
    # Columna de valores para el grafico (index 1-based)
    value_col_idx = {'unidades': 3, 'rentabilidad': 6, 'precio': 5}[value_col]
    n = len(rows_data)

    chart = BarChart()
    chart.type = 'bar'   # barras horizontales
    chart.grouping = 'clustered'
    chart.title = chart_title
    chart.y_axis.title = 'Producto'
    chart.x_axis.title = value_label
    chart.style = 10
    chart.width = 22
    chart.height = 14

    # Datos del grafico (columna de valor)
    data = Reference(ws, min_col=value_col_idx, min_row=1, max_row=1 + n)
    cats = Reference(ws, min_col=2, min_row=2, max_row=1 + n)  # descripcion
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = COLOR_HEADER

    # Ubicar el grafico debajo de la tabla
    ws.add_chart(chart, 'A%d' % (last_row + 3))


# ---------------------------------------------------------------------------
# HOJA 5 — Margen Negativo
# ---------------------------------------------------------------------------

def _hoja_margen_negativo(wb, df_neg):
    ws = wb.create_sheet('Margen Negativo')
    ws.sheet_view.showGridLines = False

    if df_neg.empty:
        ws['A1'] = 'No hay productos con margen negativo.'
        ws['A1'].font = Font(name='Arial', bold=True, color=COLOR_VERDE, size=11)
        return

    headers = ['Codigo', 'Descripcion', 'Unidades', 'Costo ($)',
               'Precio ($)', 'Rentabilidad ($)', 'Margen (%)']
    fmt_map = {3: FMT_DECIMAL, 4: FMT_MONEDA, 5: FMT_MONEDA,
               6: FMT_MONEDA, 7: FMT_PORCENTAJE}

    rows_data = [
        (row['codigo'], row['descripcion'], row['unidades'],
         row['costo'], row['precio'], row['rentabilidad'], row['margen'])
        for _, row in df_neg.iterrows()
    ]
    _write_table(ws, headers, rows_data, fmt_map=fmt_map, neg_col=7)

    # Colorear toda la fila en rojo claro
    ROJO_CLARO = 'FDECEA'
    for i in range(len(rows_data)):
        for c in range(1, 8):
            ws.cell(row=2 + i, column=c).fill = _fill(ROJO_CLARO)

    _autofit(ws)


# ---------------------------------------------------------------------------
# HOJA 6 — Analisis Pareto
# ---------------------------------------------------------------------------

def _hoja_pareto(wb, pareto_df, pareto_stats):
    ws = wb.create_sheet('Analisis Pareto')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    if pareto_df.empty:
        ws['A1'] = 'Sin datos.'
        return

    # Info del pareto
    n80 = pareto_stats.get('n_productos_80pct', 0)
    pct = pareto_stats.get('pct_productos_para_80', 0)
    total = pareto_stats.get('total_productos', 0)
    ws.merge_cells('A1:F1')
    ws['A1'] = ('Analisis Pareto: %d productos (%.1f%% del total) '
                'representan el 80%% de la facturacion') % (n80, pct)
    ws['A1'].font = Font(name='Arial', bold=True, size=11, color=COLOR_HEADER)
    ws['A1'].fill = _fill(COLOR_TITULO_BG)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    headers = ['#', 'Codigo', 'Descripcion', 'Precio ($)',
               'Precio Acumulado ($)', 'Participacion Acum. (%)']
    fmt_map = {4: FMT_MONEDA, 5: FMT_MONEDA, 6: FMT_PORCENTAJE}

    rows_data = []
    for i, (_, row) in enumerate(pareto_df.iterrows(), 1):
        rows_data.append((
            i,
            row['codigo'],
            row['descripcion'],
            row['precio'],
            row['precio_acum'],
            row['participacion_acum'],
        ))

    _write_table(ws, headers, rows_data, start_row=2, fmt_map=fmt_map)

    # Marcar fila de corte 80%
    n80_row = 2 + n80  # fila donde se alcanza el 80%
    if 2 < n80_row <= 2 + len(rows_data):
        for c in range(1, 7):
            ws.cell(row=n80_row, column=c).fill = _fill('D5F5E3')
            ws.cell(row=n80_row, column=c).font = Font(name='Arial', bold=True, size=9,
                                                        color=COLOR_VERDE)

    _autofit(ws)


# ---------------------------------------------------------------------------
# Funcion principal
# ---------------------------------------------------------------------------

def generar_excel(analisis, metadata):
    """
    Genera el Excel completo en memoria y retorna bytes.
    analisis: dict retornado por analyzer.analizar()
    metadata: dict con fecha_desde, fecha_hasta, sucursal
    """
    wb = Workbook()

    resumen = analisis['resumen']
    resumen['sucursal'] = metadata.get('sucursal', '')
    resumen['fecha_desde'] = metadata.get('fecha_desde', '')
    resumen['fecha_hasta'] = metadata.get('fecha_hasta', '')

    _hoja_resumen(wb, resumen, metadata)
    _hoja_datos(wb, analisis['df_completo'])
    _hoja_top(wb, 'Top Vendidos', analisis['top_unidades'],
              'unidades', 'Unidades Vendidas', 'Top 15 — Productos mas Vendidos')
    _hoja_top(wb, 'Top Rentabilidad', analisis['top_rentabilidad'],
              'rentabilidad', 'Rentabilidad ($)', 'Top 15 — Mayor Rentabilidad')
    _hoja_margen_negativo(wb, analisis['margen_negativo'])
    _hoja_pareto(wb, analisis['pareto_df'], analisis['pareto_stats'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
